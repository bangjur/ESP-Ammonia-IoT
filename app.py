from flask import Flask, render_template, jsonify, Response
from flask_mqtt import Mqtt
from flask_socketio import SocketIO
import sqlite3
import json
import csv
from datetime import datetime, timedelta
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, timedelta

app = Flask(__name__)

# Add subscription tracking
mqtt_subscription_active = False

# Add after other global variables:
last_sequence = None

# MQTT Settings
app.config['MQTT_BROKER_URL'] = 'broker.emqx.io'
app.config['MQTT_BROKER_PORT'] = 1883
app.config['MQTT_USERNAME'] = ''  # public broker doesn't need credentials
app.config['MQTT_PASSWORD'] = ''
app.config['MQTT_CLIENT_ID'] = 'AmmoniacServer080704'  # Change 123 to random numbers

# Initialize MQTT and SocketIO
mqtt = Mqtt(app)
socketio = SocketIO(app)

def init_db():
    """CREATES A FRESH DATABASE AND TABLES ON EVERY START"""
    conn = sqlite3.connect('sensors.db')
    c = conn.cursor()
    
    # Drop existing table if it exists
    c.execute('DROP TABLE IF EXISTS readings')
    
    # Create fresh table
    c.execute('''CREATE TABLE readings
                 (timestamp TEXT, sensor_id INTEGER, ammonia REAL)''')
    
    conn.commit()
    conn.close()
    print("Database reset completed - starting fresh!")

@mqtt.on_connect()
def handle_connect(client, userdata, flags, rc):
    """SUBSCRIBES TO ALL SENSOR TOPICS WHEN CONNECTED (ONLY ONCE)"""
    global mqtt_subscription_active
    if not mqtt_subscription_active:
        mqtt.subscribe('amoniac/sensor/#')
        mqtt_subscription_active = True
        print("Initially subscribed to MQTT topics")

@mqtt.on_message()
def handle_mqtt_message(client, userdata, message):
    """HANDLES INCOMING MQTT MESSAGES"""
    global last_sequence
    
    try:
        print(f"Received message on topic: {message.topic}")
        data = json.loads(message.payload.decode())
        print(f"Message data: {data}")
        
        # Check sequence number to avoid duplicates
        if last_sequence is not None and data.get('sequence', 0) <= last_sequence:
            print(f"Skipping duplicate or old message (sequence: {data.get('sequence')})")
            return
            
        last_sequence = data.get('sequence', 0)
        current_time = datetime.now().isoformat()
        
        # Save all readings to database
        conn = sqlite3.connect('sensors.db')
        c = conn.cursor()
        
        if 'readings' in data:  # New combined format
            for reading in data['readings']:
                c.execute('INSERT INTO readings VALUES (?, ?, ?)',
                         (current_time, reading['sensor_id'], reading['ammonia']))
                # Emit individual updates for real-time display
                socketio.emit('new_reading', reading)
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        print(f"Error processing message: {e}")
        print(f"Message payload: {message.payload}")

@app.route('/')
def index():
    """SERVES THE MAIN PAGE"""
    return render_template('index.html')

@app.route('/current_data')
def get_current_data():
    """RETURNS THE LATEST READINGS FOR EACH SENSOR"""
    conn = sqlite3.connect('sensors.db')
    c = conn.cursor()
    c.execute('''
        SELECT sensor_id, ammonia, timestamp
        FROM readings
        WHERE (sensor_id, timestamp) IN 
            (SELECT sensor_id, MAX(timestamp) 
             FROM readings 
             GROUP BY sensor_id)
    ''')
    readings = c.fetchall()
    conn.close()
    
    return jsonify([{
        'sensor_id': r[0],
        'ammonia': r[1],
        'timestamp': r[2]
    } for r in readings])

@app.route('/historical_data')
def get_historical_data():
    """RETURNS LAST 50 READINGS FOR EACH SENSOR"""
    conn = sqlite3.connect('sensors.db')
    c = conn.cursor()
    c.execute('''
        SELECT sensor_id, ammonia, timestamp
        FROM (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY sensor_id ORDER BY timestamp DESC) as rn
            FROM readings
        )
        WHERE rn <= 50
        ORDER BY timestamp ASC
    ''')
    readings = c.fetchall()
    conn.close()
    
    return jsonify([{
        'sensor_id': r[0],
        'ammonia': r[1],
        'timestamp': r[2]
    } for r in readings])

@app.route('/download_excel')
def download_excel():
    """GENERATES AN EXCEL FILE OF ALL READINGS WITH PROPER FORMATTING AND TOTALS"""
    conn = sqlite3.connect('sensors.db')
    c = conn.cursor()
    
    # Modified query to group readings within 5-second windows
    c.execute('''
        WITH TimeWindows AS (
            -- First, round timestamps to 5-second intervals
            SELECT 
                strftime('%Y-%m-%d %H:%M:', timestamp) || 
                (cast(strftime('%S', timestamp) as integer) / 5 * 5) as window_start,
                sensor_id,
                ammonia
            FROM readings
            WHERE timestamp IS NOT NULL
        ),
        GroupedData AS (
            -- Then group by these 5-second windows
            SELECT 
                window_start,
                MAX(CASE WHEN sensor_id = 1 THEN ammonia ELSE 0 END) as trash_can_a,
                MAX(CASE WHEN sensor_id = 2 THEN ammonia ELSE 0 END) as trash_can_b
            FROM TimeWindows
            GROUP BY window_start
        )
        SELECT 
            window_start as timestamp,
            trash_can_a,
            trash_can_b,
            (trash_can_a + trash_can_b) as total_ppm
        FROM GroupedData
        ORDER BY timestamp DESC
    ''')
    
    rows = c.fetchall()
    conn.close()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ammonia Readings"
    
    # Add headers with formatting
    headers = ['Timestamp', 'Trash Can A (PPM)', 'Trash Can B (PPM)', 'Total PPM']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Add data
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row[0])  # Timestamp is already formatted
        ws.cell(row=row_idx, column=2, value=round(row[1], 2))
        ws.cell(row=row_idx, column=3, value=round(row[2], 2))
        ws.cell(row=row_idx, column=4, value=round(row[3], 2))
        
        # Color coding for the total
        if row[3] > 20:  # High level
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        elif row[3] > 10:  # Medium level
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFEDCC", end_color="FFEDCC", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Add summary at the bottom
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value="Average Values")
    ws.cell(row=summary_row, column=2, value=f"=AVERAGE(B2:B{len(rows)+1})")
    ws.cell(row=summary_row, column=3, value=f"=AVERAGE(C2:C{len(rows)+1})")
    ws.cell(row=summary_row, column=4, value=f"=AVERAGE(D2:D{len(rows)+1})")
    
    # Format summary row
    for col in range(1, 5):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Create Excel file in memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return Response(
        excel_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 
            f"attachment; filename=ammonia_readings_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        }
    )

if __name__ == '__main__':
    init_db()
    socketio.run(app, debug=True)
